import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
from typing import Optional
from loguru import logger
from src.backtester import resolve_sector_member_stock_ids

# The report contains the four core sheets plus the sector-constituent appendix;
# the observe-only appendix is added when its derived data is available.
NEW_GAINER_GRADES = ("A級新起漲", "B級早期點火", "C級個股事件")
CONTINUED_GRADE = "續漲訊號"

UNCALIBRATED_WARNING = (
    "【特別警語】本系統之評分權重及訊號門檻均為先驗設計值，未經台股歷史實證回測校準"
    "(SPEC_ADDENDUM B-1)。報告內容僅供研究對照，不構成任何交易投資建議。"
)


class ReportGenerator:
    def __init__(self, output_dir: str = "C:/Workspace_CN/taiwan_moneyflow_rotation/outputs/daily"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        self.fonts = ["Microsoft JhengHei", "Microsoft YaHei", "Arial Unicode MS", "Arial"]

    def _get_active_font(self, size: int = 11, bold: bool = False, color: str = "000000") -> Font:
        return Font(name=self.fonts[0], size=size, bold=bold, color=color)

    def generate_excel_report(self,
                              trade_date: str,
                              df_sectors: pd.DataFrame,
                              df_stocks: pd.DataFrame,
                              dq_score: float,
                              dq_status: str,
                              dq_issues: list,
                              mapping_coverage_pct: Optional[float] = None,
                              sector_confidence: Optional[str] = None,
                              stock_confidence: Optional[str] = None,
                              observe_rankings: Optional[dict] = None) -> str:
        """
        Produces MoneyFlow_Rotation_YYYY-MM-DD.xlsx with the four core sheets,
        the sector-constituent appendix, and (when available) the observe-only
        appendix. Every sheet carries a data-quality-and-caveats column or block
        so a reader never mistakes a low-confidence/degraded result for a fully-
        verified one.
        """
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border_side = Side(border_style="thin", color="D3D3D3")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        self._build_dashboard_sheet(
            wb, trade_date, df_sectors, dq_score, dq_status, dq_issues,
            mapping_coverage_pct, sector_confidence, stock_confidence,
            navy_fill, light_gray_fill, thin_border,
        )
        self._build_new_gainer_sheet(wb, df_sectors, navy_fill, thin_border)
        self._build_continued_momentum_sheet(wb, df_sectors, navy_fill, thin_border)
        self._build_stock_priority_sheet(wb, df_stocks, mapping_coverage_pct, navy_fill, thin_border)
        try:
            self._build_sector_constituents_sheet(wb, df_sectors, df_stocks, navy_fill, thin_border)
        except Exception as exc:
            # This is an informational appendix; retain the established report if it fails.
            logger.exception(f"Sector constituents sheet skipped (non-fatal): {exc}")
        if observe_rankings is not None:
            try:
                self._build_observe_rankings_sheet(wb, observe_rankings, navy_fill, thin_border)
            except Exception as exc:
                # This is an observe-only appendix; retain the established report if it fails.
                logger.exception(f"Observe rankings sheet skipped (non-fatal): {exc}")

        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        filename = f"MoneyFlow_Rotation_{trade_date}.xlsx"
        full_path = os.path.join(self.output_dir, filename)
        wb.save(full_path)
        logger.info(f"Excel report saved successfully to {full_path}")
        return full_path

    def _build_sector_constituents_sheet(self, wb, df_sectors, df_stocks, navy_fill, thin_border):
        """Show scored stocks for the signalled sectors in source-row order.

        Membership is resolved through the same authoritative rule used by the
        backtester: primary sectors use ``primary_sector`` and theme sectors use
        any of ``theme_1``/``theme_2``/``theme_3``. Missing values remain blank.
        """
        ws = wb.create_sheet(title="族群成分股")
        ws.sheet_view.showGridLines = True
        ws["A1"] = (
            "本表只是把族群裡的股票攤開給你看，不是推薦名單。\n"
            "排序用的個股分數,2026-08-13 的診斷顯示它在『隔天開盤才買得到』的條件下"
            "沒有經過驗證的選股能力,請只當作瀏覽順序。"
        )
        ws["A1"].font = self._get_active_font(bold=True, color="FF0000", size=11)
        ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")

        headers = [
            "族群名稱", "訊號等級", "股票代號", "股票名稱", "收盤", "當日漲跌%",
            "5日漲跌%", "10日漲跌%", "相對成交量(20日)", "外資買賣超", "投信買賣超",
            "個股分數", "角色", "資料可信度", "處置註記",
        ]
        display_columns = [
            "sector_name", "signal_type", "stock_id", "stock_name", "close", "daily_return",
            "return_5d", "return_10d", "relative_volume_20d", "foreign_net_buy",
            "investment_trust_net_buy", "stock_score", "stock_role", "score_confidence",
            "disposition_flag",
        ]

        signal_sectors = pd.DataFrame()
        if df_sectors is not None and not df_sectors.empty and "signal_type" in df_sectors.columns:
            signal_sectors = df_sectors[
                df_sectors["signal_type"].isin(NEW_GAINER_GRADES + (CONTINUED_GRADE,))
            ]

        row = 3
        if signal_sectors.empty:
            ws.cell(row=row, column=1, value="今日無訊號族群,無成分股可列").font = self._get_active_font()
            return

        for _, sector_row in signal_sectors.iterrows():
            sector_name = sector_row.get("sector_name")
            signal_type = sector_row.get("signal_type")
            sector_type = sector_row.get("sector_type", "primary") if "sector_type" in signal_sectors.columns else "primary"
            if pd.isna(sector_type) or not str(sector_type).strip():
                sector_type = "primary"

            member_ids = []
            if df_stocks is not None and not df_stocks.empty and "stock_id" in df_stocks.columns:
                member_ids = resolve_sector_member_stock_ids(
                    str(sector_name), str(sector_type), df_stocks
                )
            member_id_set = set(member_ids)
            members = pd.DataFrame()
            if member_id_set:
                members = df_stocks[df_stocks["stock_id"].astype(str).isin(member_id_set)].copy()
            if "stock_score" in members.columns:
                members = members.sort_values("stock_score", ascending=False, na_position="last")
            members = members.head(20)

            title = f"{sector_name}｜{signal_type}｜成分股共 {len(member_id_set)} 檔,以下列出前 20 檔"
            ws.cell(row=row, column=1, value=title).font = self._get_active_font(bold=True, color="1F497D")
            row += 1
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num, value=header)
                cell.font = self._get_active_font(bold=True, color="FFFFFF")
                cell.fill = navy_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            row += 1

            for _, stock_row in members.iterrows():
                values = [stock_row.get(column) for column in display_columns]
                for col_num, value in enumerate(values, 1):
                    if pd.isna(value):
                        value = None
                    cell = ws.cell(row=row, column=col_num, value=value)
                    cell.font = self._get_active_font()
                    cell.border = thin_border
                    if col_num == 3:
                        cell.number_format = "@"
                    elif col_num in (5, 12):
                        cell.number_format = "0.0"
                    elif col_num in (6, 7, 8):
                        cell.number_format = "0.0%"
                row += 1
            row += 1

    def _build_observe_rankings_sheet(self, wb, observe_rankings, navy_fill, thin_border):
        """Add an observe-only appendix without changing the established four sheets."""
        ws = wb.create_sheet(title="觀察-前300與法人")
        ws.sheet_view.showGridLines = True
        ws["A1"] = "觀察-前300與法人（僅觀察，未接入訊號評級）"
        ws["A1"].font = self._get_active_font(bold=True, color="FF0000", size=12)
        ws["A2"] = "平手規則：指標優先，其次成交值由高到低，再依股票代號升冪；法人資料單位為股數。"
        ws["A2"].font = self._get_active_font(size=10, color="595959")

        row = 4
        headings = {
            "top_gainers": "漲幅前300：族群聚合",
            "top_losers": "跌幅前300：族群聚合",
            "top_turnover": "成交值前300：族群聚合",
        }
        for key, title in headings.items():
            ws.cell(row=row, column=1, value=title).font = self._get_active_font(bold=True, color="1F497D")
            row += 1
            headers = ["族群", "上榜檔數", "族群總檔數", "上榜比例", "上榜成交值", "前300成交值占比"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self._get_active_font(bold=True, color="FFFFFF")
                cell.fill = navy_fill
                cell.border = thin_border
            row += 1
            for item in observe_rankings["rankings"][key]["top5_sectors"]:
                values = [item["sector_name"], item["selected_stock_count"], item["sector_stock_count"],
                          item["selected_stock_ratio"], item["selected_turnover"], item.get("turnover_share_of_top300")]
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value).border = thin_border
                ws.cell(row=row, column=4).number_format = "0.0%"
                ws.cell(row=row, column=6).number_format = "0.0%"
                row += 1
            row += 1

        ws.cell(row=row, column=1, value="三大法人族群聚合（淨買超前5）").font = self._get_active_font(bold=True, color="1F497D")
        row += 1
        headers = ["族群", "外資", "投信", "自營商", "合計淨買賣超"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self._get_active_font(bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.border = thin_border
        row += 1
        for item in observe_rankings["institutional_by_sector"]["net_buy_top5"]:
            for col, value in enumerate([item["sector_name"], item["foreign_net_buy"], item["investment_trust_net_buy"], item["dealer_net_buy"], item["net_buy_total"]], 1):
                ws.cell(row=row, column=col, value=value).border = thin_border
            row += 1

    # ------------------------------------------------------------------
    # Sheet 1: Dashboard
    # ------------------------------------------------------------------
    def _build_dashboard_sheet(self, wb, trade_date, df_sectors, dq_score, dq_status, dq_issues,
                                mapping_coverage_pct, sector_confidence, stock_confidence,
                                navy_fill, light_gray_fill, thin_border):
        ws = wb.create_sheet(title="Dashboard")
        ws.sheet_view.showGridLines = True

        ws["A1"] = "台股主流族群與資金輪動追蹤系統"
        ws["A1"].font = self._get_active_font(size=16, bold=True, color="1F497D")

        ws["A2"] = f"交易日期: {trade_date} | 執行時間: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = self._get_active_font(size=11, bold=False, color="595959")

        ws["A4"] = UNCALIBRATED_WARNING
        ws["A4"].font = self._get_active_font(size=10, bold=True, color="FF0000")

        row = 6
        ws.cell(row=row, column=1, value="資料品質與警語 (Data Quality & Caveats)").font = self._get_active_font(size=12, bold=True)
        row += 1

        ws.cell(row=row, column=1, value="資料品質得分").font = self._get_active_font(bold=True)
        ws.cell(row=row, column=1).fill = light_gray_fill
        ws.cell(row=row, column=2, value=(dq_score if dq_score is not None else np.nan))
        ws.cell(row=row, column=2).number_format = "0.00"
        row += 1

        ws.cell(row=row, column=1, value="資料品質狀態").font = self._get_active_font(bold=True)
        ws.cell(row=row, column=1).fill = light_gray_fill
        ws.cell(row=row, column=2, value=dq_status or "UNKNOWN")
        row += 1

        coverage_pct = mapping_coverage_pct if mapping_coverage_pct is not None else np.nan
        ws.cell(row=row, column=1, value="產業映射覆蓋率").font = self._get_active_font(bold=True)
        ws.cell(row=row, column=1).fill = light_gray_fill
        cov_cell = ws.cell(row=row, column=2, value=coverage_pct)
        cov_cell.number_format = "0.0%"
        row += 1

        ws.cell(row=row, column=1, value="族群評分信心等級").font = self._get_active_font(bold=True)
        ws.cell(row=row, column=1).fill = light_gray_fill
        ws.cell(row=row, column=2, value=sector_confidence or "N/A")
        row += 1

        ws.cell(row=row, column=1, value="個股評分信心等級").font = self._get_active_font(bold=True)
        ws.cell(row=row, column=1).fill = light_gray_fill
        ws.cell(row=row, column=2, value=stock_confidence or "N/A")
        row += 2

        # Explicit, unmissable low-mapping-coverage warning (task brief requirement:
        # "Dashboard 必須誠實顯示覆蓋率與『絕大多數股票未分類→族群訊號不可用/降級』警語")
        if mapping_coverage_pct is not None and mapping_coverage_pct < 0.80:
            ws.cell(row=row, column=1,
                    value=(f"【重大警語】產業映射覆蓋率僅 {mapping_coverage_pct:.2%}，"
                           "絕大多數股票尚未分類至產業族群。族群層新起漲/續漲訊號"
                           "覆蓋範圍嚴重受限、可信度大幅降級，僅少數已映射個股所屬族群"
                           "可視為有效訊號來源；個股層排序（不依賴產業分類）仍可正常使用。")
                    ).font = self._get_active_font(bold=True, color="FF0000", size=11)
            row += 2

        ws.cell(row=row, column=1, value="資料品質異常與警告清單:").font = self._get_active_font(bold=True, size=11)
        row += 1
        if not dq_issues:
            ws.cell(row=row, column=1, value="無異常。資料品質良好。").font = self._get_active_font()
            row += 1
        else:
            for issue in dq_issues:
                ws.cell(row=row, column=1, value=f"- {issue}").font = self._get_active_font(color="9C0006")
                row += 1
        row += 1

        # Top-10 sector score table (Dashboard requirement: 前10族群分數表)
        ws.cell(row=row, column=1, value="前10族群分數摘要").font = self._get_active_font(size=12, bold=True)
        row += 1
        headers = ["族群名稱", "族群類型", "今日總分", "上漲廣度", "成交占比", "訊號等級", "資料信心", "警語"]
        header_row = row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = self._get_active_font(bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        if df_sectors is not None and not df_sectors.empty and "score" in df_sectors.columns:
            top10 = df_sectors.sort_values("score", ascending=False).head(10)
            for _, r in top10.iterrows():
                caveat = "族群映射覆蓋率低，訊號僅供參考" if (mapping_coverage_pct is not None and mapping_coverage_pct < 0.80) else ""
                values = [
                    r.get("sector_name"), r.get("sector_type", ""), r.get("score"),
                    r.get("breadth"), r.get("volume_share"),
                    r.get("signal_type", "無訊號"), r.get("signal_data_confidence", r.get("score_confidence", "N/A")),
                    caveat,
                ]
                for c_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=c_idx, value=val)
                    cell.font = self._get_active_font()
                    cell.border = thin_border
                    if c_idx == 3:
                        cell.number_format = "0.0"
                    elif c_idx in (4, 5):
                        cell.number_format = "0.0%"
                row += 1
        else:
            ws.cell(row=row, column=1, value="今日無族群分數資料可供摘要。").font = self._get_active_font()

    # ------------------------------------------------------------------
    # Sheet 2: New Gainer Sectors (新起漲族群)
    # ------------------------------------------------------------------
    def _build_new_gainer_sheet(self, wb, df_sectors, navy_fill, thin_border):
        ws = wb.create_sheet(title="新起漲族群")
        ws.sheet_view.showGridLines = True

        ws["A1"] = "新起漲族群 (New Gainer Sectors) — 訊號門檻未經台股實證校準，僅供研究參考"
        ws["A1"].font = self._get_active_font(bold=True, color="FF0000", size=11)

        df_new = pd.DataFrame()
        if df_sectors is not None and not df_sectors.empty and "signal_type" in df_sectors.columns:
            df_new = df_sectors[df_sectors["signal_type"].isin(NEW_GAINER_GRADES)].copy()

        headers = ["族群名稱", "訊號等級", "今日總分", "觸發原因", "未通過條件", "失效條件",
                   "前五大個股", "資料可信度"]
        header_row = 3
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = self._get_active_font(bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")

        r_idx = header_row + 1
        if not df_new.empty:
            df_new = df_new.sort_values("score", ascending=False)
            for _, row in df_new.iterrows():
                top5 = row.get("top5_stocks", "N/A")
                values = [
                    row.get("sector_name"), row.get("signal_type"), row.get("score"),
                    row.get("signal_reason", row.get("conditions_passed", "")),
                    row.get("conditions_failed", "(無)"),
                    row.get("invalidation_condition", ""),
                    top5,
                    row.get("signal_data_confidence", "N/A"),
                ]
                for c_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = self._get_active_font()
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if c_idx == 3:
                        cell.number_format = "0.0"
                r_idx += 1
        else:
            ws.cell(row=r_idx, column=1, value="今日無符合條件之新起漲族群。").font = self._get_active_font()
            r_idx += 1

        r_idx += 1
        ws.cell(row=r_idx, column=1, value="資料品質與警語：門檻均為 PLACEHOLDER - UNCALIBRATED（未經回測校準），"
                                            "族群訊號僅涵蓋已完成產業映射之個股，未分類個股不計入族群訊號判斷。"
                ).font = self._get_active_font(size=10, color="9C0006")

    # ------------------------------------------------------------------
    # Sheet 3: Continued Momentum Sectors (續漲族群)
    # ------------------------------------------------------------------
    def _build_continued_momentum_sheet(self, wb, df_sectors, navy_fill, thin_border):
        ws = wb.create_sheet(title="續漲族群")
        ws.sheet_view.showGridLines = True

        ws["A1"] = "續漲族群 (Continued Momentum Sectors) — 訊號門檻未經台股實證校準，僅供研究參考"
        ws["A1"].font = self._get_active_font(bold=True, color="FF0000", size=11)

        df_cont = pd.DataFrame()
        if df_sectors is not None and not df_sectors.empty and "signal_type" in df_sectors.columns:
            df_cont = df_sectors[df_sectors["signal_type"] == CONTINUED_GRADE].copy()

        headers = ["族群名稱", "訊號等級", "今日總分", "觸發原因", "未通過條件", "失效條件",
                   "前五大個股", "資料可信度"]
        header_row = 3
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = self._get_active_font(bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")

        r_idx = header_row + 1
        if not df_cont.empty:
            df_cont = df_cont.sort_values("score", ascending=False)
            for _, row in df_cont.iterrows():
                top5 = row.get("top5_stocks", "N/A")
                values = [
                    row.get("sector_name"), row.get("signal_type"), row.get("score"),
                    row.get("signal_reason", ""),
                    row.get("conditions_failed", "(無)"),
                    row.get("invalidation_condition", ""),
                    top5,
                    row.get("signal_data_confidence", "N/A"),
                ]
                for c_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = self._get_active_font()
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if c_idx == 3:
                        cell.number_format = "0.0"
                r_idx += 1
        else:
            ws.cell(row=r_idx, column=1, value="今日無符合條件之續漲族群。").font = self._get_active_font()
            r_idx += 1

        r_idx += 1
        ws.cell(row=r_idx, column=1, value="資料品質與警語：門檻均為 PLACEHOLDER - UNCALIBRATED（未經回測校準），"
                                            "族群訊號僅涵蓋已完成產業映射之個股，未分類個股不計入族群訊號判斷。"
                ).font = self._get_active_font(size=10, color="9C0006")

    # ------------------------------------------------------------------
    # Sheet 4: Stock Observation Priority (個股優先排序)
    # ------------------------------------------------------------------
    def _build_stock_priority_sheet(self, wb, df_stocks, mapping_coverage_pct, navy_fill, thin_border):
        ws = wb.create_sheet(title="個股優先排序")
        ws.sheet_view.showGridLines = True

        ws["A1"] = "個股優先排序 (Stock Observation Priority) — 個股排序不依賴產業分類，即使族群未分類仍可用"
        ws["A1"].font = self._get_active_font(bold=True, color="1F497D", size=11)

        headers = ["股票代號", "股票名稱", "所屬族群", "個股角色", "總分", "訊號原因",
                   "風險原因", "資料可信度", "處置/注意"]
        header_row = 3
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = self._get_active_font(bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")

        r_idx = header_row + 1
        if df_stocks is not None and not df_stocks.empty and "stock_score" in df_stocks.columns:
            df_display = df_stocks.sort_values("stock_score", ascending=False, na_position="last").head(50).copy()
            for _, row in df_display.iterrows():
                primary_sector = row.get("primary_sector", "待分類")
                risk_reason = self._build_stock_risk_reason(row)
                signal_reason = self._build_stock_signal_reason(row)
                confidence = row.get("score_confidence", "N/A")
                if primary_sector in ("待分類", "未分類", None) or (isinstance(primary_sector, float) and pd.isna(primary_sector)):
                    if confidence == "FULL":
                        confidence = "DEGRADED"
                disposition_flag = row.get("disposition_flag", "N/A(未查核)")

                values = [
                    row.get("stock_id"), row.get("stock_name"), primary_sector,
                    row.get("stock_role", ""), row.get("stock_score"),
                    signal_reason, risk_reason, confidence, disposition_flag,
                ]
                for c_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = self._get_active_font()
                    cell.border = thin_border
                    if c_idx == 1:
                        cell.number_format = "@"
                    elif c_idx == 5:
                        cell.number_format = "0.0"
                    if c_idx == 9 and disposition_flag in ("處置股", "注意股"):
                        cell.font = self._get_active_font(bold=True, color="9C0006")
                r_idx += 1
        else:
            ws.cell(row=r_idx, column=1, value="今日無符合條件之個股排序。").font = self._get_active_font()
            r_idx += 1

        r_idx += 1
        cov_text = f"{mapping_coverage_pct:.2%}" if mapping_coverage_pct is not None else "N/A"
        ws.cell(row=r_idx, column=1,
                value=(f"資料品質與警語：產業映射覆蓋率 {cov_text}；「所屬族群」顯示為"
                       "「待分類」的個股尚未完成產業對照，其總分與角色判斷不依賴族群分數，"
                       "但族群相關的訊號原因/風險原因會相應降級標註。「處置/注意」欄="
                       "「N/A(未查核)」代表當日未執行處置/注意股查核（非「確認正常」）。")
                ).font = self._get_active_font(size=10, color="9C0006")

    @staticmethod
    def _build_stock_signal_reason(row) -> str:
        parts = []
        role = row.get("stock_role")
        if role:
            parts.append(f"角色:{role}")
        sector_score = row.get("sector_score")
        if pd.notna(sector_score):
            parts.append(f"所屬族群分數{sector_score:.1f}")
        rank_improvement = row.get("rank_improvement")
        if pd.notna(rank_improvement) and rank_improvement is not None:
            try:
                if float(rank_improvement) > 0:
                    parts.append("排名改善")
            except (TypeError, ValueError):
                pass
        return "; ".join(parts) if parts else "資料不足，無法判定訊號原因"

    @staticmethod
    def _build_stock_risk_reason(row) -> str:
        risks = []
        turnover = row.get("turnover")
        if pd.notna(turnover) and turnover is not None and turnover < 1_000_000:
            risks.append("日成交額過低")
        foreign_net_buy = row.get("foreign_net_buy")
        if pd.notna(foreign_net_buy) and foreign_net_buy is not None and foreign_net_buy < 0:
            risks.append("外資賣超")
        primary_sector = row.get("primary_sector")
        if primary_sector in ("待分類", "未分類", None) or (isinstance(primary_sector, float) and pd.isna(primary_sector)):
            risks.append("尚未完成產業分類，族群層風險無法評估")
        return "; ".join(risks) if risks else "無明顯風險標記"
