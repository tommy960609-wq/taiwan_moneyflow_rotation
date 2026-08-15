import sys
import os

import pandas as pd
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../../")))

from src.leaderboard_loader import (
    discover_leaderboard_files, load_one_leaderboard, load_all_leaderboards,
    flag_limit_up_proxy, LIMIT_UP_PROXY_THRESHOLD_PCT,
)


def _write_report_xlsx(path, rows):
    """rows: list of (rank, stock_id, stock_name, return_pct, turnover_million)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["排名", "代號", "名稱", "漲跌幅", "成交額(百萬)"])
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def test_discover_leaderboard_files_sorted_by_date(tmp_path):
    _write_report_xlsx(tmp_path / "Report_20260716.xlsx", [(1, "2330", "台積電", 5.0, 100.0)])
    _write_report_xlsx(tmp_path / "Report_20260515.xlsx", [(1, "2330", "台積電", 5.0, 100.0)])
    files = discover_leaderboard_files(str(tmp_path))
    assert len(files) == 2
    assert os.path.basename(files[0]) == "Report_20260515.xlsx"
    assert os.path.basename(files[1]) == "Report_20260716.xlsx"


def test_load_one_leaderboard_normalizes_columns_and_trade_date(tmp_path):
    path = tmp_path / "Report_20260716.xlsx"
    _write_report_xlsx(path, [
        (1, "2330", "台積電", 5.0, 100.0),
        (2, "1101", "台泥", 9.8, 50.0),
    ])
    df = load_one_leaderboard(str(path))
    assert list(df.columns) == ["trade_date", "rank", "stock_id", "stock_name",
                                 "return_pct", "turnover_million_twd"]
    assert (df["trade_date"] == "2026-07-16").all()
    assert df.iloc[0]["stock_id"] == "2330"
    assert df.iloc[1]["return_pct"] == 9.8


def test_load_one_leaderboard_wrong_column_count_fails_closed(tmp_path):
    path = tmp_path / "Report_20260716.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["排名", "代號", "名稱"])  # only 3 columns, malformed
    ws.append([1, "2330", "台積電"])
    wb.save(path)
    df = load_one_leaderboard(str(path))
    assert df.empty


def test_load_one_leaderboard_bad_filename_returns_empty(tmp_path):
    path = tmp_path / "NotAReport.xlsx"
    _write_report_xlsx(path, [(1, "2330", "台積電", 5.0, 100.0)])
    df = load_one_leaderboard(str(path))
    assert df.empty


def test_load_all_leaderboards_stacks_multiple_files_and_skips_bad_ones(tmp_path):
    _write_report_xlsx(tmp_path / "Report_20260515.xlsx", [(1, "2330", "台積電", 5.0, 100.0)])
    _write_report_xlsx(tmp_path / "Report_20260716.xlsx", [(1, "1101", "台泥", 3.0, 20.0)])
    # A malformed file that should be skipped, not crash the whole load.
    wb = openpyxl.Workbook()
    wb.active.append(["only", "two"])
    wb.save(tmp_path / "Report_20260601.xlsx")

    df = load_all_leaderboards(str(tmp_path))
    assert len(df) == 2
    assert set(df["trade_date"]) == {"2026-05-15", "2026-07-16"}


def test_load_all_leaderboards_empty_dir_returns_empty_with_columns(tmp_path):
    df = load_all_leaderboards(str(tmp_path))
    assert df.empty
    assert "trade_date" in df.columns


def test_flag_limit_up_proxy_threshold():
    df = pd.DataFrame({
        "return_pct": [10.0, 9.5, 9.49, 0.0, -5.0],
    })
    out = flag_limit_up_proxy(df)
    assert out["limit_up_proxy"].tolist() == [True, True, False, False, False]
    assert LIMIT_UP_PROXY_THRESHOLD_PCT == 9.5
