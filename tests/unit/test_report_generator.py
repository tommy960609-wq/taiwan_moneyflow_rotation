import sys
import os
import tempfile
import openpyxl
import pandas as pd
import numpy as np

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../../")))

from src.report_generator import ReportGenerator

EXPECTED_SHEET_NAMES = ["Dashboard", "新起漲族群", "續漲族群", "個股優先排序", "族群成分股"]


def _mock_sectors():
    return pd.DataFrame([
        {
            "sector_name": "SectorA", "sector_type": "primary", "may_double_count": False,
            "score": 78.0, "breadth": 0.65, "volume_share": 0.20, "signal_type": "A級新起漲",
            "signal_reason": "十項條件全數通過，族群性新起漲成立",
            "conditions_passed": "rule1: ...", "conditions_failed": "(無)",
            "invalidation_condition": "失效條件：族群總分回落至55以下...",
            "signal_data_confidence": "FULL", "top5_stocks": "台積電(2330)、鴻海(2317)",
        },
        {
            "sector_name": "SectorB", "sector_type": "primary", "may_double_count": False,
            "score": 66.0, "breadth": 0.5, "volume_share": 0.12, "signal_type": "續漲訊號",
            "signal_reason": "資金結構與強度維持穩定上揚，未見明確破壞跡象",
            "conditions_passed": "rule1: ...", "conditions_failed": "(無)",
            "invalidation_condition": "失效條件：族群總分跌破65...",
            "signal_data_confidence": "DEGRADED", "top5_stocks": "聯發科(2454)",
        },
        {
            "sector_name": "SectorC", "sector_type": "theme", "may_double_count": True,
            "score": 40.0, "breadth": 0.2, "volume_share": 0.05, "signal_type": "無訊號",
            "signal_reason": "未達新起漲或續漲門檻",
            "conditions_passed": "(無)", "conditions_failed": "rule1: ...",
            "invalidation_condition": "無有效訊號，無失效條件適用",
            "signal_data_confidence": "LOW", "top5_stocks": "N/A",
        },
    ])


def _mock_stocks():
    return pd.DataFrame([
        {
            "stock_id": "2330", "stock_name": "台積電", "primary_sector": "半導體",
            "stock_role": "領先龍頭", "stock_score": 88.0, "sector_score": 78.0,
            "rank_improvement": 5, "turnover": 5_000_000_000.0, "foreign_net_buy": 1000.0,
            "score_confidence": "FULL",
        },
        {
            "stock_id": "9999", "stock_name": "未分類個股", "primary_sector": "待分類",
            "stock_role": "低位階補漲股", "stock_score": 40.0, "sector_score": np.nan,
            "rank_improvement": -2, "turnover": 500_000.0, "foreign_net_buy": -200.0,
            "score_confidence": "FULL",
        },
    ])


def _generate(tmp_dir, mapping_coverage_pct=0.9, sectors=None, stocks=None, observe_rankings=None):
    gen = ReportGenerator(output_dir=tmp_dir)
    return gen.generate_excel_report(
        trade_date="2026-07-16",
        df_sectors=_mock_sectors() if sectors is None else sectors,
        df_stocks=_mock_stocks() if stocks is None else stocks,
        dq_score=92.5,
        dq_status="WARNING",
        dq_issues=["Sample issue for test"],
        mapping_coverage_pct=mapping_coverage_pct,
        sector_confidence="DEGRADED",
        stock_confidence="FULL",
        observe_rankings=observe_rankings,
    )


class TestCoreSheetsAndConstituents:
    def test_exactly_five_sheets_with_sector_constituents(self):
        """The report has the four core sheets plus the constituent appendix."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            assert wb.sheetnames == EXPECTED_SHEET_NAMES, f"Expected exactly {EXPECTED_SHEET_NAMES}, got {wb.sheetnames}"
            assert len(wb.sheetnames) == 5


class TestDashboardSheet:
    def test_dashboard_contains_quality_score_and_mapping_coverage(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir, mapping_coverage_pct=0.05)
            wb = openpyxl.load_workbook(path)
            ws = wb["Dashboard"]
            all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None)
            assert "資料品質" in all_text
            assert "映射覆蓋率" in all_text
            # Low coverage (5%) must trigger the explicit "most stocks unclassified" warning
            assert "絕大多數股票尚未分類" in all_text
            assert "PLACEHOLDER" not in all_text or "未經台股歷史實證回測校準" in all_text

    def test_dashboard_uncalibrated_warning_always_present(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["Dashboard"]
            all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None)
            assert "未經台股歷史實證回測校準" in all_text

    def test_dashboard_top10_sector_table_present(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["Dashboard"]
            all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None)
            assert "SectorA" in all_text
            assert "族群名稱" in all_text


class TestNewGainerSheet:
    def test_required_columns_present(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["新起漲族群"]
            header_row = [c.value for c in ws[3]]
            for required in ["族群名稱", "訊號等級", "今日總分", "觸發原因", "未通過條件", "失效條件", "前五大個股", "資料可信度"]:
                assert required in header_row, f"Missing required column {required}"

    def test_only_new_gainer_grades_included_not_continued_or_none(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["新起漲族群"]
            rows_text = [str(c.value) for r in ws.iter_rows(min_row=4) for c in r if c.value]
            assert "SectorA" in " ".join(rows_text)
            assert "SectorB" not in " ".join(rows_text)  # SectorB is 續漲訊號, not new-gainer
            assert "SectorC" not in " ".join(rows_text)  # SectorC is 無訊號

    def test_data_quality_caveat_present(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["新起漲族群"]
            all_text = " ".join(str(c.value) for r in ws.iter_rows() for c in r if c.value is not None)
            assert "資料品質" in all_text or "PLACEHOLDER" in all_text


class TestContinuedMomentumSheet:
    def test_required_columns_present(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["續漲族群"]
            header_row = [c.value for c in ws[3]]
            for required in ["族群名稱", "訊號等級", "今日總分", "觸發原因", "未通過條件", "失效條件", "前五大個股", "資料可信度"]:
                assert required in header_row

    def test_only_continued_grade_included(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["續漲族群"]
            rows_text = " ".join(str(c.value) for r in ws.iter_rows(min_row=4) for c in r if c.value)
            assert "SectorB" in rows_text
            assert "SectorA" not in rows_text


class TestStockPrioritySheet:
    def test_required_columns_present(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["個股優先排序"]
            header_row = [c.value for c in ws[3]]
            for required in ["股票代號", "股票名稱", "所屬族群", "個股角色", "總分", "訊號原因", "風險原因", "資料可信度"]:
                assert required in header_row

    def test_stock_data_present_and_non_empty(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["個股優先排序"]
            rows_text = " ".join(str(c.value) for r in ws.iter_rows(min_row=4) for c in r if c.value)
            assert "2330" in rows_text
            assert "9999" in rows_text

    def test_unclassified_stock_gets_downgraded_confidence_and_risk_note(self):
        """
        A stock whose primary_sector is still 待分類 (unclassified) must show a risk
        note about being unclassified and must not silently claim FULL confidence for
        sector-dependent context, per the task brief's "not allowed to pretend
        unclassified stocks have sector signals" requirement.
        """
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["個股優先排序"]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row[0] == "9999":
                    row_text = " ".join(str(v) for v in row if v is not None)
                    assert "尚未完成產業分類" in row_text
                    break
            else:
                raise AssertionError("Row for stock 9999 not found")


class TestNumberFormatting:
    def test_percentage_columns_use_percent_format(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir, mapping_coverage_pct=0.42)
            wb = openpyxl.load_workbook(path)
            ws = wb["Dashboard"]
            found_pct_format = any(
                cell.number_format == "0.0%"
                for row in ws.iter_rows() for cell in row
            )
            assert found_pct_format, "Expected at least one 0.0% formatted cell on Dashboard (mapping coverage)"


class TestM7DispositionColumn:
    """M7: 個股優先排序 sheet gains a '處置/注意' column reflecting
    src/disposition_fetcher.py's real same-day snapshot."""

    def test_disposition_column_header_present(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["個股優先排序"]
            header_row = [c.value for c in ws[3]]
            assert "處置/注意" in header_row

    def test_missing_disposition_flag_defaults_to_not_checked(self):
        """Stocks in the mock frame have no disposition_flag column at all (as would
        happen for a historical backfill day with no disposition fetch run) -- must
        show the honest 'not checked' sentinel, never blank or a fabricated '正常'."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            ws = wb["個股優先排序"]
            rows_text = " ".join(str(c.value) for r in ws.iter_rows(min_row=4) for c in r if c.value)
            assert "N/A(未查核)" in rows_text

    def test_real_disposition_flag_displayed(self):
        gen = ReportGenerator(output_dir=tempfile.mkdtemp())
        df_stocks = _mock_stocks().copy()
        df_stocks["disposition_flag"] = ["處置股", "正常"]
        path = gen.generate_excel_report(
            trade_date="2026-07-16", df_sectors=_mock_sectors(), df_stocks=df_stocks,
            dq_score=92.5, dq_status="WARNING", dq_issues=[],
            mapping_coverage_pct=0.9, sector_confidence="DEGRADED", stock_confidence="FULL",
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["個股優先排序"]
        found_2330_disposition = False
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] == "2330":
                assert "處置股" in row
                found_2330_disposition = True
        assert found_2330_disposition


def _mock_constituent_stocks(count=3):
    rows = []
    for idx in range(count):
        rows.append({
            "stock_id": str(9000 + idx), "stock_name": f"SectorA-{idx}",
            "primary_sector": "SectorA", "stock_score": float(100 - idx),
            "close": 10.0 + idx, "daily_return": 0.01 * idx,
            "return_5d": 0.02 * idx, "return_10d": 0.03 * idx,
            "relative_volume_20d": 1.0 + idx, "foreign_net_buy": 100 + idx,
            "investment_trust_net_buy": 50 + idx, "stock_role": "領先龍頭",
            "score_confidence": "FULL", "disposition_flag": "正常",
        })
    rows.append({
        "stock_id": "8000", "stock_name": "SectorB-0", "primary_sector": "SectorB",
        "stock_score": 88.0, "close": 20.0, "daily_return": 0.02,
        "return_5d": 0.04, "return_10d": 0.06, "relative_volume_20d": 1.2,
        "foreign_net_buy": 200, "investment_trust_net_buy": 80,
        "stock_role": "次族群龍頭", "score_confidence": "DEGRADED",
        "disposition_flag": "N/A(未查核)",
    })
    return pd.DataFrame(rows)


def _mock_observe_rankings():
    empty_rankings = {
        "top_gainers": {"top5_sectors": []},
        "top_losers": {"top5_sectors": []},
        "top_turnover": {"top5_sectors": []},
    }
    return {
        "rankings": empty_rankings,
        "institutional_by_sector": {"net_buy_top5": []},
    }


class TestSectorConstituentsSheet:
    def test_signal_scope_score_order_and_cap(self):
        sectors = _mock_sectors()
        stocks = _mock_constituent_stocks(count=25)
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir, sectors=sectors, stocks=stocks)
            wb = openpyxl.load_workbook(path)
            ws = wb["族群成分股"]
            values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
            assert "SectorC" not in " ".join(str(value) for value in values)
            title_row = next(row for row in range(1, ws.max_row + 1) if str(ws.cell(row, 1).value).startswith("SectorA"))
            assert "成分股共 25 檔,以下列出前 20 檔" in ws.cell(title_row, 1).value
            header_row = title_row + 1
            first_stock_row = header_row + 1
            scores = [ws.cell(first_stock_row + offset, 12).value for offset in range(20)]
            assert scores == sorted(scores, reverse=True)
            assert all(ws.cell(first_stock_row + offset, 3).value is not None for offset in range(20))

    def test_a1_plain_language_warnings_present(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir)
            wb = openpyxl.load_workbook(path)
            warning = wb["族群成分股"]["A1"].value
            assert "本表只是把族群裡的股票攤開給你看，不是推薦名單。" in warning
            assert "沒有經過驗證的選股能力,請只當作瀏覽順序。" in warning

    def test_no_signal_day_has_prompt_row(self):
        sectors = _mock_sectors().copy()
        sectors["signal_type"] = "無訊號"
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = _generate(tmp_dir, sectors=sectors)
            wb = openpyxl.load_workbook(path)
            ws = wb["族群成分股"]
            assert ws["A3"].value == "今日無訊號族群,無成分股可列"

    def test_constituent_failure_keeps_other_five_sheets(self, monkeypatch):
        gen = ReportGenerator(output_dir=tempfile.mkdtemp())

        def fail_builder(*args, **kwargs):
            raise RuntimeError("intentional test failure")

        monkeypatch.setattr(gen, "_build_sector_constituents_sheet", fail_builder)
        path = gen.generate_excel_report(
            trade_date="2026-07-16", df_sectors=_mock_sectors(), df_stocks=_mock_stocks(),
            dq_score=92.5, dq_status="WARNING", dq_issues=[], mapping_coverage_pct=0.9,
            sector_confidence="DEGRADED", stock_confidence="FULL",
            observe_rankings=_mock_observe_rankings(),
        )
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == [
            "Dashboard", "新起漲族群", "續漲族群", "個股優先排序", "觀察-前300與法人"
        ]
