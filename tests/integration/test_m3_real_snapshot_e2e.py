import sys
import os
import json
import shutil
import tempfile
import openpyxl
import pandas as pd
import pytest
from unittest.mock import patch

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../../")))

from scripts.run_daily import run_pipeline
from src.config_manager import ConfigManager

RAW_SAMPLES_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../loop/evidence/raw_samples"))
REAL_MAPPING = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../data/reference/stock_industry_mapping.xlsx"))
TRADE_DATE = "2026-07-16"


def _load_payload(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("payload") if isinstance(data, dict) and "payload" in data else data


@pytest.mark.slow
def test_real_snapshot_2026_07_16_end_to_end_produces_real_daily_excel(tmp_path):
    """
    Milestone 3 end-to-end requirement: run the full daily pipeline (clean -> map ->
    features -> score -> signal -> Excel) against the REAL cached 2026-07-16 TWSE+TPEx
    OHLCV/institutional/margin snapshots from loop/evidence/raw_samples (no network
    calls, no synthetic data). Data and outputs are isolated beneath pytest's temp
    directory, while the resulting MoneyFlow_Rotation_2026-07-16.xlsx is still checked
    for all 4 SPEC_ADDENDUM B-4 sheets and internal consistency.

    Input data provenance (verified against loop/evidence/raw_samples metadata):
      - TWSE/TPEx OHLCV: both dated 1150716 (ROC) = 2026-07-16, 1371 + 10012 raw rows.
      - TPEx institutional/margin: dated 1150716 = 2026-07-16 (matches OHLCV date).
      - TWSE institutional (T86): fetched for query-date 20260717 (T86 has no
        multi-date payload; the cached sample's query param is the day *after* the
        target trade date). This is a pre-existing evidence-snapshot limitation, not
        something this test fabricates -- it is explicitly disclosed in the Milestone
        3 Acceptance Report rather than silently treated as a 07-16 match.
      - TWSE margin (MI_MARGN): position-keyed dict rows, no per-row date field.

    Both `data_dir` and `output_dir` are redirected to isolated temp paths so this test
    never mutates the project's real data fixtures or outputs.

    Known environment-coupling finding (disclosed in Milestone_3_Acceptance_Report.md):
    `scripts.run_daily.load_excel_leaderboard` (pre-existing M1 code, out of M3's
    modify scope per the "don't touch M1/M2 accepted module behavior" rule) globs
    `C:/Workspace_CN/Quant-Agent/**/Report_<date>.xlsx` -- an external sibling project
    directory. On this development machine a real leaderboard file for 2026-07-16
    happens to exist there, and reconciling against it surfaces a real, pre-existing
    return-basis mismatch: `reconcile_with_leaderboard`/`calculate_ranks` compute
    daily_return as (close-open)/open (an intraday-range proxy), whereas the
    leaderboard's 漲跌幅 column is the standard prev-close-to-close percentage change.
    That mismatch degrades the Data Quality Score enough to BLOCK the pipeline (a
    correct fail-closed reaction to the deviation count, not a bug in this M3 delivery).
    Because that leaderboard file is an incidental, non-reproducible artifact of this
    particular machine (a clean checkout elsewhere would not have it, and it is not a
    documented required M3 input), this test patches `load_excel_leaderboard` to return
    empty -- exactly mirroring how `tests/integration/test_run_daily.py` already patches
    file-existence checks for hermetic isolation -- so the real-snapshot pipeline run
    reflects the reproducible, spec-mandated data flow (OHLCV+institutional+margin),
    not an accidental cross-project side effect.
    """
    with tempfile.TemporaryDirectory(dir=str(tmp_path)) as temp_dir:
        temp_data_dir = os.path.join(temp_dir, "data")
        temp_output_dir = os.path.join(temp_dir, "outputs")
        os.makedirs(os.path.join(temp_data_dir, "reference"), exist_ok=True)
        os.makedirs(os.path.join(temp_data_dir, "raw/ohlcv"), exist_ok=True)
        os.makedirs(os.path.join(temp_data_dir, "raw/institutional"), exist_ok=True)
        os.makedirs(os.path.join(temp_data_dir, "raw/margin"), exist_ok=True)
        os.makedirs(os.path.join(temp_data_dir, "processed"), exist_ok=True)
        os.makedirs(temp_output_dir, exist_ok=True)

        shutil.copy(REAL_MAPPING, os.path.join(temp_data_dir, "reference/stock_industry_mapping.xlsx"))

        twse_ohlcv = _load_payload(os.path.join(RAW_SAMPLES_DIR, "twse_ohlcv_sample.json"))
        tpex_ohlcv = _load_payload(os.path.join(RAW_SAMPLES_DIR, "tpex_ohlcv_sample.json"))
        with open(os.path.join(temp_data_dir, f"raw/ohlcv/twse_prices_{TRADE_DATE}.json"), "w", encoding="utf-8") as f:
            json.dump({"payload": twse_ohlcv}, f, ensure_ascii=False)
        with open(os.path.join(temp_data_dir, f"raw/ohlcv/tpex_prices_{TRADE_DATE}.json"), "w", encoding="utf-8") as f:
            json.dump({"payload": tpex_ohlcv}, f, ensure_ascii=False)

        twse_inst_full = json.load(open(os.path.join(RAW_SAMPLES_DIR, "twse_inst_sample.json"), encoding="utf-8"))["payload"]
        tpex_inst = _load_payload(os.path.join(RAW_SAMPLES_DIR, "tpex_inst_sample.json"))
        with open(os.path.join(temp_data_dir, f"raw/institutional/inst_{TRADE_DATE}.json"), "w", encoding="utf-8") as f:
            json.dump({"payload": twse_inst_full}, f, ensure_ascii=False)
        with open(os.path.join(temp_data_dir, f"raw/institutional/tpex_inst_{TRADE_DATE}.json"), "w", encoding="utf-8") as f:
            json.dump({"payload": tpex_inst}, f, ensure_ascii=False)

        twse_margin = _load_payload(os.path.join(RAW_SAMPLES_DIR, "twse_margin_sample.json"))
        tpex_margin = _load_payload(os.path.join(RAW_SAMPLES_DIR, "tpex_margin_sample.json"))
        with open(os.path.join(temp_data_dir, f"raw/margin/margin_{TRADE_DATE}.json"), "w", encoding="utf-8") as f:
            json.dump({"payload": twse_margin}, f, ensure_ascii=False)
        with open(os.path.join(temp_data_dir, f"raw/margin/tpex_margin_{TRADE_DATE}.json"), "w", encoding="utf-8") as f:
            json.dump({"payload": tpex_margin}, f, ensure_ascii=False)

        def mock_config_get(self, key, default=None):
            if key == "system.data_dir":
                return temp_data_dir
            if key == "system.output_dir":
                return temp_output_dir
            if key == "system.run_mode":
                return "production"
            defaults = ConfigManager().get_defaults()
            keys = key.split(".")
            val = defaults
            for k in keys:
                if isinstance(val, dict) and k in val:
                    val = val[k]
                else:
                    return default
            return val

        with patch("src.config_manager.ConfigManager.get", mock_config_get), \
             patch("scripts.run_daily.load_excel_leaderboard", return_value=pd.DataFrame()):
            run_pipeline(TRADE_DATE)

        expected_excel = os.path.join(temp_output_dir, "daily", f"MoneyFlow_Rotation_{TRADE_DATE}.xlsx")
        assert os.path.commonpath([expected_excel, str(tmp_path)]) == str(tmp_path)
        assert os.path.exists(expected_excel), f"Real-snapshot daily Excel not found at {expected_excel}"
        assert os.path.getsize(expected_excel) > 0

        wb = openpyxl.load_workbook(expected_excel)
        assert wb.sheetnames == [
            "Dashboard", "新起漲族群", "續漲族群", "個股優先排序", "族群成分股",
            "觀察-前300與法人",
        ], wb.sheetnames

        # Dashboard must always honestly show the DQ block and the actual mapping
        # coverage percentage. Whether the "most stocks unclassified" warning appears
        # depends on the CURRENT real coverage (M5a official mapping import raised
        # coverage from the old 0.41% baseline to ~98% -- see
        # docs/Milestone_5a_Acceptance_Report.md), so this assertion must be
        # conditional on the actual number, not hardcoded to the old low-coverage state.
        ws_dash = wb["Dashboard"]
        dash_text = " ".join(str(c.value) for r in ws_dash.iter_rows() for c in r if c.value is not None)
        assert "資料品質" in dash_text
        assert "映射覆蓋率" in dash_text

        audit_path_for_coverage = os.path.join(temp_output_dir, "logs", f"audit_{TRADE_DATE}.json")
        with open(audit_path_for_coverage, encoding="utf-8") as f:
            audit_for_coverage = json.load(f)
        coverage_pct = audit_for_coverage["row_counts"].get("mapping_coverage_pct")
        assert coverage_pct is not None, "audit JSON must record mapping_coverage_pct"

        if coverage_pct < 0.80:
            assert "絕大多數股票尚未分類" in dash_text, "Low mapping coverage warning must be present, not hidden"
        else:
            assert "絕大多數股票尚未分類" not in dash_text, (
                f"Mapping coverage is {coverage_pct:.2%} (>=80%); the low-coverage "
                "warning must not appear once coverage is healthy"
            )
            coverage_str_variants = [f"{coverage_pct:.1%}", f"{coverage_pct * 100:.1f}%", f"{coverage_pct:.0%}"]
            assert any(v in dash_text for v in coverage_str_variants) or str(round(coverage_pct * 100)) in dash_text, (
                f"Dashboard must display the actual coverage number ({coverage_pct:.2%}) once healthy"
            )

        # Stock priority sheet must be non-empty and have real stock IDs (individual-
        # stock ranking must work even though sector mapping coverage is low).
        ws_stock = wb["個股優先排序"]
        stock_rows = [r for r in ws_stock.iter_rows(min_row=4, values_only=True) if r[0]]
        assert len(stock_rows) > 0, "Stock priority sheet must be non-empty for the real full-market snapshot"

        # Log file and audit JSON (task brief requirement 4) must exist for this run.
        log_path = os.path.join(temp_output_dir, "logs", f"run_{TRADE_DATE}.log")
        audit_path = os.path.join(temp_output_dir, "logs", f"audit_{TRADE_DATE}.json")
        assert os.path.exists(log_path), f"Expected run log at {log_path}"
        assert os.path.exists(audit_path), f"Expected audit summary at {audit_path}"

        with open(audit_path, encoding="utf-8") as f:
            audit = json.load(f)
        assert audit["status"] == "SUCCESS"
        assert audit["row_counts"]["prices_twse"] > 0
        assert audit["row_counts"]["prices_tpex"] > 0
        assert len(audit["output_files"]) > 0

        print(
            f"Real-snapshot M3 E2E OK: {len(stock_rows)} stocks in priority sheet, "
            f"audit={audit_path}, excel={expected_excel} ({os.path.getsize(expected_excel)} bytes)"
        )
