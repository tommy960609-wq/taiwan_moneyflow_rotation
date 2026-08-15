import openpyxl

def check_sheets():
    path = "C:/Workspace_CN/taiwan_moneyflow_rotation/outputs/daily/MoneyFlow_Rotation_2026-07-16.xlsx"
    wb = openpyxl.load_workbook(path)
    print(f"Loaded Excel: {path}")
    print(f"Sheets count: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")
    
    # Check Dashboard content
    ws_dash = wb["Dashboard"]
    print(f"A1 Cell (Title): {ws_dash['A1'].value}")
    print(f"A4 Cell (Warning): {ws_dash['A4'].value}")
    print(f"B7 Cell (Quality Score): {ws_dash['B7'].value}")
    print(f"B8 Cell (Quality Status): {ws_dash['B8'].value}")
    
    # Check Priority sheet
    ws_stock = wb["Stock Observation Priority"]
    print(f"A1 Cell (First Header): {ws_stock['A1'].value}")
    print(f"A2 Cell (First stock ID): {ws_stock['A2'].value}")
    print(f"B2 Cell (First stock Name): {ws_stock['B2'].value}")
    print(f"E2 Cell (First stock Role): {ws_stock['E2'].value}")

if __name__ == "__main__":
    check_sheets()
