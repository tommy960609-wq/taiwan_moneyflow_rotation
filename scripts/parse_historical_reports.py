import glob
import openpyxl

def try_decode(s):
    if not isinstance(s, str):
        return s
    encodings = ['latin1', 'utf-8', 'cp950', 'gbk']
    decodings = ['cp950', 'utf-8', 'big5', 'gbk']
    
    for enc in encodings:
        for dec in decodings:
            try:
                candidate = s.encode(enc).decode(dec)
                # Check if it looks like common Chinese characters
                if any(c in candidate for c in ["股", "代", "名", "收", "量", "成", "強", "信"]):
                    return candidate
            except:
                pass
    return s

def parse_reports():
    files = glob.glob('C:/Workspace_CN/Quant-Agent/**/Report_20260716.xlsx', recursive=True)
    if not files:
        print("Historical reports not found.")
        return
        
    path = files[0]
    wb = openpyxl.load_workbook(path)
    print(f"File: {path}")
    print("Sheets:", wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        for r_idx in range(1, 10):
            row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, 15)]
            decoded_vals = []
            for v in row_vals:
                if v is not None:
                    dec = try_decode(v)
                    decoded_vals.append(dec)
                else:
                    decoded_vals.append(None)
            print(f"Row {r_idx}: {decoded_vals}")

if __name__ == "__main__":
    parse_reports()
